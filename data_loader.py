"""Lectura y parseo del Excel de servicio becario.

El Excel tiene una hoja de semestre (ej. "Propuesta de 2027-1") con tres
bloques apilados verticalmente:

1. Lista de eventos (encabezados en la fila 1).
2. Tabla de becas: "% Beca" -> "Horas requeridas".
3. Becarios activos: Nombre, % Beca, Rol, Semestre, Eval. 360.

Todo se detecta por el TEXTO de los encabezados, nunca por posición fija,
para tolerar que muevan columnas o agreguen filas entre bloques.
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Utilidades de texto
# ---------------------------------------------------------------------------

def normalize_text(value) -> str:
    """Minúsculas, sin acentos y con espacios colapsados. '' si viene vacío."""
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


SPANISH_MONTHS = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12,
}


def parse_date_like(value, default_year: int | None = None):
    """Convierte una celda de fecha a `date`.

    Acepta datetime/date nativos y textos como "22-jul-2026", "3-7 ago 2026"
    o "12-14 nov" (rangos: se toma el primer día; sin año: `default_year`).
    Devuelve None si no se puede interpretar.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text:
        return None

    month = None
    for token, num in SPANISH_MONTHS.items():
        if re.search(rf"\b{token}", text):
            month = num
            break
    if month is None:
        return None

    day_match = re.search(r"\b(\d{1,2})\b", text)
    year_match = re.search(r"\b(\d{4})\b", text)
    if day_match is None:
        return None
    year = int(year_match.group(1)) if year_match else default_year
    if year is None:
        return None
    try:
        return date(year, month, int(day_match.group(1)))
    except ValueError:
        return None


def parse_hours_value(value):
    """Convierte 'Horas Contabilizables' a float: acepta 4.5, "6.5h", "3H"."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().lower().replace(",", ".")
    match = re.fullmatch(r"([\d.]+)\s*h?(?:rs?|oras)?\.?", text)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return None
    return None


# --- Parseo de rangos de hora ("17:00 a 19:00", "9:30AM-2:00PM", "9am a 1pm")

_TIME_TOKEN = re.compile(
    r"(\d{1,2}):(\d{2})\s*([ap])?\.?\s*m?\.?|(\d{1,2})\s*([ap])\.?\s*m\.?",
    re.IGNORECASE,
)


def _extract_time_tokens(text: str):
    tokens = []
    for m in _TIME_TOKEN.finditer(text):
        if m.group(1) is not None:
            hour, minute = int(m.group(1)), int(m.group(2))
            meridiem = m.group(3).lower() if m.group(3) else None
        else:
            hour, minute = int(m.group(4)), 0
            meridiem = m.group(5).lower()
        if hour <= 24 and minute < 60:
            tokens.append((hour, minute, meridiem))
    return tokens


def _to_decimal(hour: int, minute: int, meridiem) -> float:
    if meridiem == "p" and hour < 12:
        hour += 12
    elif meridiem == "a" and hour == 12:
        hour = 0
    return hour + minute / 60


def parse_time_range(text) -> float | None:
    """Duración en horas del primer rango de hora encontrado en el texto."""
    if text is None:
        return None
    tokens = _extract_time_tokens(str(text))
    if len(tokens) < 2:
        return None
    (h1, m1, mer1), (h2, m2, mer2) = tokens[0], tokens[1]

    # Inferir am/pm faltante a partir del otro extremo del rango.
    if mer1 is None and mer2 is not None and h1 <= 12:
        mer1 = mer2 if _to_decimal(h1, m1, mer2) <= _to_decimal(h2, m2, mer2) else ("a" if mer2 == "p" else "p")
    if mer2 is None and mer1 is not None and h2 <= 12:
        mer2 = mer1 if _to_decimal(h1, m1, mer1) <= _to_decimal(h2, m2, mer1) else ("p" if mer1 == "a" else "a")

    start, end = _to_decimal(h1, m1, mer1), _to_decimal(h2, m2, mer2)
    if end <= start and end + 12 <= 24:
        end += 12  # "11:00 a 1:00" sin am/pm: el fin cruza mediodía
    duration = end - start
    if 0 < duration <= 16:
        return round(duration, 2)
    return None


# ---------------------------------------------------------------------------
# Detección de hojas y columnas
# ---------------------------------------------------------------------------

# nombre lógico -> función que reconoce el encabezado normalizado
_HEADER_MATCHERS = {
    "evento": lambda h: h == "evento",
    "hora": lambda h: h == "hora",
    "horas": lambda h: "horas contab" in h or "valor del evento" in h,
    "lugar": lambda h: h == "lugar",
    "becarios": lambda h: "becario" in h,
    "encargado": lambda h: h == "encargado",
}


def _map_header_columns(ws) -> dict:
    """Mapea nombre lógico -> índice de columna leyendo la fila 1."""
    columns: dict[str, int] = {}
    date_candidates = []
    for col in range(1, ws.max_column + 1):
        header = normalize_text(ws.cell(row=1, column=col).value)
        if not header:
            continue
        for key, matcher in _HEADER_MATCHERS.items():
            if key not in columns and matcher(header):
                columns[key] = col
        if "fecha" in header:
            date_candidates.append(col)

    # Hay dos columnas "FECHA"/"Fecha": nos quedamos con la que tenga más
    # fechas reales (celdas datetime).
    best_col, best_count = None, -1
    for col in date_candidates:
        count = sum(
            1 for row in range(2, min(ws.max_row, 400) + 1)
            if isinstance(ws.cell(row=row, column=col).value, datetime)
        )
        if count > best_count:
            best_col, best_count = col, count
    if best_col is not None:
        columns["fecha"] = best_col
    return columns


def sheet_scores(file_bytes: bytes) -> list[tuple[str, int]]:
    """[(hoja, puntaje)] ordenado: qué tanto se parece cada hoja al formato
    de eventos. Puntúa columnas clave presentes + bloques de becas/becarios."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    scores = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = {normalize_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1), [])}
        score = 0
        for key, matcher in _HEADER_MATCHERS.items():
            if any(matcher(h) for h in headers if h):
                score += 1
        if score >= 2:  # solo buscar bloques donde ya parece hoja de eventos
            block_hits = 0
            for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                for v in row:
                    norm = normalize_text(v)
                    if norm in ("tabla de becas", "becarios activos"):
                        block_hits += 1
            score += 2 * min(block_hits, 2)
        scores.append((name, score))
    wb.close()
    scores.sort(key=lambda item: -item[1])
    return scores


# ---------------------------------------------------------------------------
# Parseo de los tres bloques
# ---------------------------------------------------------------------------

@dataclass
class ParsedSheet:
    sheet_name: str
    events: pd.DataFrame = field(default_factory=pd.DataFrame)
    tabla_becas: pd.DataFrame = field(default_factory=pd.DataFrame)
    becarios: pd.DataFrame = field(default_factory=pd.DataFrame)
    warnings: list = field(default_factory=list)


def _find_marker_row(ws, texts: tuple, min_row: int = 2) -> int | None:
    for row in range(min_row, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 15) + 1):
            if normalize_text(ws.cell(row=row, column=col).value) in texts:
                return row
    return None


def _normalize_pct(value) -> float | None:
    """% de beca a número 0-100 (acepta 0.3 formateado como % o 30 plano)."""
    if value is None:
        return None
    if isinstance(value, str):
        match = re.search(r"[\d.]+", value.replace(",", "."))
        if not match:
            return None
        value = float(match.group(0))
    value = float(value)
    if value <= 1:
        value *= 100
    return round(value, 2)


def _parse_events(ws, columns: dict, stop_row: int) -> tuple[pd.DataFrame, list]:
    warnings = []
    col_evento = columns.get("evento")
    col_fecha = columns.get("fecha")

    # Año por defecto para fechas escritas sin año ("12-14 nov").
    default_year = None
    if col_fecha:
        years = [
            ws.cell(row=r, column=col_fecha).value.year
            for r in range(2, stop_row)
            if isinstance(ws.cell(row=r, column=col_fecha).value, datetime)
        ]
        if years:
            default_year = max(set(years), key=years.count)

    rows = []
    for r in range(2, stop_row):
        evento = ws.cell(row=r, column=col_evento).value if col_evento else None
        if evento is None or not str(evento).strip():
            continue  # separadores de mes y filas vacías

        def cell(key):
            col = columns.get(key)
            return ws.cell(row=r, column=col).value if col else None

        hora_texto = cell("hora")
        horas = parse_hours_value(cell("horas"))
        horas_estimadas = False
        if horas is None:
            estimada = parse_time_range(hora_texto)
            if estimada is not None:
                horas, horas_estimadas = estimada, True

        rows.append({
            "fila": r,
            "evento": re.sub(r"\s+", " ", str(evento)).strip(),
            "fecha": parse_date_like(cell("fecha"), default_year),
            "hora_texto": str(hora_texto).strip() if hora_texto is not None else "",
            "horas": horas if horas is not None else 0.0,
            "horas_estimadas": horas_estimadas,
            "horas_pendientes": horas is None,
            "lugar": str(cell("lugar")).strip() if cell("lugar") is not None else "",
            "becarios_raw": str(cell("becarios")).strip() if cell("becarios") is not None else "",
            "encargado": str(cell("encargado")).strip() if cell("encargado") is not None else "",
        })

    if not rows:
        warnings.append("No se encontraron eventos en la hoja.")
    return pd.DataFrame(rows), warnings


def _parse_tabla_becas(ws) -> pd.DataFrame:
    for row in range(1, ws.max_row + 1):
        cols = {}
        for col in range(1, min(ws.max_column, 15) + 1):
            norm = normalize_text(ws.cell(row=row, column=col).value)
            if norm == "% beca":
                cols["pct"] = col
            elif "horas requeridas" in norm:
                cols["horas"] = col
        if "pct" in cols and "horas" in cols:
            rows = []
            r = row + 1
            while r <= ws.max_row:
                pct = _normalize_pct(ws.cell(row=r, column=cols["pct"]).value)
                horas = parse_hours_value(ws.cell(row=r, column=cols["horas"]).value)
                if pct is None or horas is None:
                    break
                rows.append({"pct_beca": pct, "horas_requeridas": horas})
                r += 1
            return pd.DataFrame(rows)
    return pd.DataFrame()


def _parse_becarios(ws) -> pd.DataFrame:
    for row in range(1, ws.max_row + 1):
        cols = {}
        for col in range(1, min(ws.max_column, 15) + 1):
            norm = normalize_text(ws.cell(row=row, column=col).value)
            if norm == "nombre":
                cols["nombre"] = col
            elif norm == "% beca":
                cols["pct"] = col
            elif norm == "rol":
                cols["rol"] = col
            elif norm == "semestre":
                cols["semestre"] = col
            elif "eval" in norm or "360" in norm:
                cols["eval"] = col
        if "nombre" in cols and "pct" in cols:
            rows = []
            r = row + 1
            while r <= ws.max_row:
                nombre = ws.cell(row=r, column=cols["nombre"]).value
                if nombre is None or not str(nombre).strip():
                    break
                eval_raw = normalize_text(ws.cell(row=r, column=cols["eval"]).value) if "eval" in cols else ""
                rows.append({
                    "nombre": re.sub(r"\s+", " ", str(nombre)).strip(),
                    "pct_beca": _normalize_pct(ws.cell(row=r, column=cols["pct"]).value),
                    "rol": str(ws.cell(row=r, column=cols["rol"]).value or "").strip() if "rol" in cols else "",
                    "semestre": str(ws.cell(row=r, column=cols["semestre"]).value or "").strip() if "semestre" in cols else "",
                    "eval_360": "Completada" if "complet" in eval_raw else "Pendiente",
                })
                r += 1
            return pd.DataFrame(rows)
    return pd.DataFrame()


def parse_sheet(file_bytes: bytes, sheet_name: str) -> ParsedSheet:
    """Parsea los tres bloques de una hoja de semestre."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name]
    result = ParsedSheet(sheet_name=sheet_name)

    columns = _map_header_columns(ws)
    missing = [k for k in ("evento", "becarios", "horas") if k not in columns]
    if missing:
        result.warnings.append(
            f"La hoja '{sheet_name}' no tiene las columnas esperadas: faltan {', '.join(missing)}. "
            "Se parseará lo que se pueda."
        )

    # Los eventos terminan donde empieza el bloque de becas / becarios.
    marker = _find_marker_row(ws, ("tabla de becas", "% beca", "becarios activos"))
    stop_row = marker if marker else ws.max_row + 1

    if "evento" in columns:
        result.events, event_warnings = _parse_events(ws, columns, stop_row)
        result.warnings.extend(event_warnings)

    result.tabla_becas = _parse_tabla_becas(ws)
    if result.tabla_becas.empty:
        result.warnings.append("No se encontró la 'Tabla de becas' (% Beca → Horas requeridas) en esta hoja.")

    result.becarios = _parse_becarios(ws)
    if result.becarios.empty:
        result.warnings.append("No se encontró la tabla de 'Becarios activos' en esta hoja.")

    wb.close()
    return result
